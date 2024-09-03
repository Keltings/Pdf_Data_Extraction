from extraction.process_pdf import process_pdf
from combined.merged import fill_missing_images, fill_missing_images_madrid, fill_industrial
from openpyxl import Workbook
import pandas as pd
import os

if __name__ == "__main__":
    main_folder_path = input('Enter Main Folder Path: ')
    
    # Get list of all subdirectories
    subdirectories = [subdir for subdir in os.listdir(main_folder_path) if os.path.isdir(os.path.join(main_folder_path, subdir))]
    
    for subdir in subdirectories:
        folder_path = os.path.join(main_folder_path, subdir)
        output_excel_path = f'output/{subdir}.xlsx'
        output_excel_madrid_path = f'output/{subdir}_Madrid.xlsx'
        output_excel_industrial_path = f'output/{subdir}_Industrial_Design.xlsx'
        output_excel = Workbook()
        output_excel_madrid = Workbook()
        output_excel_industrial = Workbook()

        # Remove the default "Sheet" created when Workbook is initialized
        default_sheet = output_excel.active
        output_excel.remove(default_sheet)

        default_sheet_madrid = output_excel_madrid.active
        output_excel_madrid.remove(default_sheet_madrid)

        default_sheet_industrial = output_excel_industrial.active
        output_excel_industrial.remove(default_sheet_industrial)

        if int(subdir) < 2012:
            for filename in os.listdir(folder_path):
                if filename.endswith(".pdf"):
                    pdf_file_path = os.path.join(folder_path, filename)
                    df_text,df_madrid_text,df_industrial_text, df_image, df_madrid = process_pdf(pdf_file_path,text_after_kenya=False)
                    fill_missing_images(df_text, df_image, output_excel, sheet_name=filename)
                    fill_missing_images_madrid(df_madrid_text, df_madrid, output_excel_madrid, sheet_name=filename)
                    fill_industrial(df_industrial_text, output_excel_industrial, sheet_name=filename)

            output_excel.save(output_excel_path)
            print(f"Excel file for {subdir} created successfully.")

            output_excel_madrid.save(output_excel_madrid_path)
            print(f"Madrid file for {subdir} created successfully.")

            output_excel_industrial.save(output_excel_industrial_path)
            print(f"Industrial Design file for {subdir} created successfully.")
            
        else:
            for filename in os.listdir(folder_path):
                if filename.endswith(".pdf"):
                    pdf_file_path = os.path.join(folder_path, filename)
                    df_text,df_madrid_text,df_industrial_text, df_image, df_madrid = process_pdf(pdf_file_path,text_after_kenya=True)
                    fill_missing_images(df_text, df_image, output_excel, sheet_name=filename)
                    fill_missing_images_madrid(df_madrid_text, df_madrid, output_excel_madrid, sheet_name=filename)
                    fill_industrial(df_industrial_text, output_excel_industrial, sheet_name=filename)

            output_excel.save(output_excel_path)
            print(f"Excel file for {subdir} created successfully.")

            output_excel_madrid.save(output_excel_madrid_path)
            print(f"Madrid file for {subdir} created successfully.")

            output_excel_industrial.save(output_excel_industrial_path)
            print(f"Industrial Design file for {subdir} created successfully.")

            